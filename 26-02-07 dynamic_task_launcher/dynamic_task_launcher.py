
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
""" (Dynamic Task Launcher)
엑셀 파일에 경로 등을 입력하고, 해당 내용을 읽어 실행하는 시스템
"""

import os
import re
import sys
import subprocess
import time
from pathlib import Path
from datetime import datetime

# UI 상수 (박스·폭)
UI_WIDTH = 80
BOX_TOP = "╔" + "═" * (UI_WIDTH - 2) + "╗"
BOX_BOT = "╚" + "═" * (UI_WIDTH - 2) + "╝"
BOX_LR = "║"
BOX_SEP = "╠" + "═" * (UI_WIDTH - 2) + "╣"
BOX_THIN = "─" * UI_WIDTH
BOX_DOUBLE = "╔" + "═" * (UI_WIDTH - 2) + "╗"

# 아이콘 정의
ICONS = {
    'python': '🐍',
    'excel': '📊',
    'folder': '📁',
    'file': '📄',
    'success': '✨',
    'error': '❌',
    'warning': '⚠️',
    'clock': '🕐',
    'rocket': '🚀',
    'star': '⭐',
    'fire': '🔥',
    'target': '🎯',
    'gem': '💎',
    'lightning': '⚡'
}

try:
    from colorama import init, Fore, Style, Back
    init(autoreset=True)
except ImportError:
    print("colorama 라이브러리가 필요합니다. 설치 중...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "colorama"])
    from colorama import init, Fore, Style, Back
    init(autoreset=True)

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("openpyxl 라이브러리가 필요합니다. 설치 중...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import Workbook


# 엑셀 설정 파일명 (스크립트와 같은 폴더)
EXCEL_FILENAME = "task_config.xlsx"
EXCEL_HEADERS = ("번호", "제목", "설명", "파일경로")


def setup_console_appearance():
    """Windows: 콘솔 폰트 및 창 크기 설정 (개선 버전)"""
    if os.name != 'nt':
        return
    try:
        import ctypes
        LF_FACESIZE = 32
        STD_OUTPUT_HANDLE = -11

        class COORD(ctypes.Structure):
            _fields_ = [("X", ctypes.c_short), ("Y", ctypes.c_short)]

        class SMALL_RECT(ctypes.Structure):
            _fields_ = [
                ("Left", ctypes.c_short),
                ("Top", ctypes.c_short),
                ("Right", ctypes.c_short),
                ("Bottom", ctypes.c_short)
            ]

        class CONSOLE_SCREEN_BUFFER_INFO(ctypes.Structure):
            _fields_ = [
                ("dwSize", COORD),
                ("dwCursorPosition", COORD),
                ("wAttributes", ctypes.c_ushort),
                ("srWindow", SMALL_RECT),
                ("dwMaximumWindowSize", COORD)
            ]

        class CONSOLE_FONT_INFOEX(ctypes.Structure):
            _fields_ = [
                ("cbSize", ctypes.c_ulong),
                ("nFont", ctypes.c_ulong),
                ("dwFontSize", COORD),
                ("FontFamily", ctypes.c_uint),
                ("FontWeight", ctypes.c_uint),
                ("FaceName", ctypes.c_wchar * LF_FACESIZE),
            ]

        kernel32 = ctypes.windll.kernel32
        handle = kernel32.GetStdHandle(STD_OUTPUT_HANDLE)

        # 폰트 설정
        font = CONSOLE_FONT_INFOEX()
        font.cbSize = ctypes.sizeof(CONSOLE_FONT_INFOEX)
        font.nFont = 0
        font.dwFontSize.X = 0
        font.dwFontSize.Y = 20  # 폰트 크기 증가
        font.FontFamily = 54
        font.FontWeight = 600  # 더 굵게
        font.FaceName = "Cascadia Code"

        if not kernel32.SetCurrentConsoleFontEx(handle, ctypes.c_long(0), ctypes.pointer(font)):
            font.FaceName = "Consolas"
            font.dwFontSize.Y = 18
            kernel32.SetCurrentConsoleFontEx(handle, ctypes.c_long(0), ctypes.pointer(font))

        # 버퍼 및 윈도우 크기 설정
        buffer_size = COORD(100, 1000)
        kernel32.SetConsoleScreenBufferSize(handle, buffer_size)

        window_size = SMALL_RECT(0, 0, 99, 35)  # 더 큰 창
        kernel32.SetConsoleWindowInfo(handle, ctypes.c_long(1), ctypes.pointer(window_size))

    except Exception:
        pass


class TaskLauncher:
    def __init__(self):
        self.script_dir = Path(__file__).resolve().parent
        self.excel_path = self.script_dir / EXCEL_FILENAME
        self.tasks = []

    def get_excel_path(self):
        """엑셀 설정 파일 전체 경로"""
        return str(self.excel_path)

    def ensure_excel_template(self):
        """엑셀 파일이 없으면 헤더만 있는 템플릿 생성"""
        if not self.excel_path.exists():
            print(f"\n{Fore.LIGHTYELLOW_EX}{ICONS['warning']} 설정 파일이 없습니다. 템플릿을 생성합니다...{Style.RESET_ALL}")

            wb = Workbook()
            ws = wb.active
            ws.title = "작업목록"
            for col, header in enumerate(EXCEL_HEADERS, 1):
                ws.cell(row=1, column=col, value=header)
            ws.cell(row=2, column=1, value=1)
            ws.cell(row=2, column=2, value="예시: Python 스크립트")
            ws.cell(row=2, column=3, value="Python 파일 실행 예시")
            ws.cell(row=2, column=4, value="C:\\경로\\스크립트.py")
            ws.cell(row=3, column=1, value=2)
            ws.cell(row=3, column=2, value="예시: 엑셀 파일")
            ws.cell(row=3, column=3, value="엑셀 문서 열기 예시")
            ws.cell(row=3, column=4, value="C:\\경로\\문서.xlsx")
            wb.save(self.get_excel_path())

            self.print_loading_animation("템플릿 생성", 0.5)
            print(f"{Fore.GREEN}{ICONS['success']} 설정 파일을 생성했습니다!{Style.RESET_ALL}")
            print(f"{Fore.LIGHTGREEN_EX}  ▸ {self.excel_path}{Style.RESET_ALL}\n")
            time.sleep(0.5)

    def open_excel_for_edit(self):
        """엑셀 파일을 기본 프로그램으로 열기"""
        path = self.get_excel_path()
        if not os.path.exists(path):
            self.ensure_excel_template()
            path = self.get_excel_path()
        try:
            if os.name == 'nt':
                os.startfile(path)
            else:
                subprocess.Popen(['xdg-open', path])
        except Exception as e:
            print(f"\n{Fore.RED}{ICONS['error']} 엑셀 파일 열기 오류{Style.RESET_ALL}")
            print(f"{Fore.LIGHTRED_EX}  ▸ {str(e)}{Style.RESET_ALL}\n")

    def load_tasks_from_excel(self):
        """엑셀 파일에서 작업 목록 읽기 (제목, 설명, 파일경로)"""
        self.tasks = []
        path = self.get_excel_path()
        if not os.path.exists(path):
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 4:
                    continue
                _, name, desc, path_val = (row[0], row[1], row[2], row[3])
                name = (name or "").strip()
                desc = (desc or "").strip()
                path_val = (path_val or "").strip()
                path_val = self.clean_path(path_val)
                if not path_val:
                    continue
                self.tasks.append({
                    "name": name or "제목 없음",
                    "desc": desc or "",
                    "path": path_val
                })
            wb.close()
        except Exception as e:
            print(f"\n{Fore.RED}{ICONS['error']} 엑셀 읽기 오류{Style.RESET_ALL}")
            print(f"{Fore.LIGHTRED_EX}  ▸ {str(e)}{Style.RESET_ALL}\n")

    def clear_screen(self):
        """화면 클리어"""
        os.system('cls' if os.name == 'nt' else 'clear')

    def get_file_modified_time(self, filepath):
        """파일의 최종 수정 시간 가져오기"""
        try:
            if os.path.exists(filepath):
                timestamp = os.path.getmtime(filepath)
                return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M')
            return "파일 없음"
        except Exception:
            return "알 수 없음"

    def get_days_since_modified(self, filepath):
        """마지막 수정일로부터 경과한 일수 (오늘 0시 기준). 없으면 None"""
        try:
            if not os.path.exists(filepath):
                return None
            mtime = os.path.getmtime(filepath)
            modified_date = datetime.fromtimestamp(mtime).date()
            today = datetime.now().date()
            return (today - modified_date).days
        except Exception:
            return None

    def file_exists(self, filepath):
        """파일 존재 여부 확인"""
        return os.path.exists(filepath)

    def _box_line(self, text, color=Fore.CYAN):
        """박스 안 한 줄 (좌우 여백, 폭 맞춤)"""
        w = UI_WIDTH - 4
        t = (text[:w] + "...") if len(text) > w else text
        pad = w - len(t)
        return f"{color}{BOX_LR}{Style.RESET_ALL} {t}{' ' * max(0, pad)} {color}{BOX_LR}{Style.RESET_ALL}"

    def _visible_len(self, text):
        """ANSI 이스케이프 제외한 표시 길이"""
        return len(re.sub(r"\033\[[0-9;]*m", "", text))

    def _card_line(self, text):
        """작업 카드 내 한 줄 (박스 폭에 맞춤, ANSI 고려)"""
        w = UI_WIDTH - 6
        visible = self._visible_len(text)
        pad = max(0, w - visible)
        return f"  {text}{' ' * pad}  "

    def _center_in_box(self, text, width=UI_WIDTH - 2):
        """박스 안 중앙 정렬 텍스트 (문자 수 기준)"""
        t = text.strip()
        pad = width - len(t)
        if pad <= 0:
            return t[:width]
        return " " * (pad // 2) + t + " " * (pad - pad // 2)

    def print_animated_header(self):
        """화려한 애니메이션 헤더 출력"""
        colors = [Fore.CYAN, Fore.LIGHTCYAN_EX, Fore.LIGHTBLUE_EX, Fore.CYAN]

        # 상단 테두리 (그라디언트 효과)
        print(f"\n{Fore.CYAN}{Style.BRIGHT}╔{'═' * (UI_WIDTH - 2)}╗{Style.RESET_ALL}")

        # 제목 라인
        title = f"{ICONS['rocket']}  JP 통합업무 대시보드  {ICONS['star']}"
        centered = self._center_in_box(title, UI_WIDTH - 2)
        print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{Back.BLUE}{Fore.WHITE}{Style.BRIGHT}{centered}{Style.RESET_ALL}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")

        # 부제목
        subtitle = "✨ Task Launcher & Workspace Manager ✨"
        centered_sub = self._center_in_box(subtitle, UI_WIDTH - 2)
        print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{Back.BLUE}{Fore.LIGHTYELLOW_EX}{centered_sub}{Style.RESET_ALL}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")

        # 구분선
        print(f"{Fore.CYAN}{Style.BRIGHT}╠{'═' * (UI_WIDTH - 2)}╣{Style.RESET_ALL}\n")

    def get_file_icon(self, filepath):
        """파일 확장자에 따른 아이콘 반환"""
        ext = Path(filepath).suffix.lower()
        if ext == '.py':
            return ICONS['python']
        elif ext in ['.xlsx', '.xls', '.csv']:
            return ICONS['excel']
        elif os.path.isdir(filepath):
            return ICONS['folder']
        else:
            return ICONS['file']

    def get_status_badge(self, days):
        """경과 일수에 따른 상태 배지"""
        if days is None:
            return f"{Fore.RED}{ICONS['error']} 파일 없음{Style.RESET_ALL}"
        elif days == 0:
            return f"{Fore.GREEN}{ICONS['fire']} 오늘 수정{Style.RESET_ALL}"
        elif days <= 3:
            return f"{Fore.LIGHTGREEN_EX}{ICONS['lightning']} {days}일 전{Style.RESET_ALL}"
        elif days <= 7:
            return f"{Fore.YELLOW}{ICONS['clock']} {days}일 전{Style.RESET_ALL}"
        else:
            return f"{Fore.LIGHTBLACK_EX}{ICONS['clock']} {days}일 전{Style.RESET_ALL}"

    def display_main_menu(self):
        """메인 화면 출력 (박스형 대시보드) - 개선된 버전"""
        self.clear_screen()

        # 화려한 헤더
        self.print_animated_header()

        # 메뉴 리스트
        if not self.tasks:
            # 작업이 없을 때
            print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{'  ' * (UI_WIDTH // 2 - 1)}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")
            msg = f"{ICONS['warning']} 등록된 작업이 없습니다"
            print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{self._center_in_box(msg, UI_WIDTH - 2)}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")
            hint = "[E]를 눌러 엑셀에서 작업을 추가하세요"
            print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{Fore.YELLOW}{self._center_in_box(hint, UI_WIDTH - 2)}{Style.RESET_ALL}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")
            print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{'  ' * (UI_WIDTH // 2 - 1)}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")
        else:
            # 작업 목록 표시
            for idx, task in enumerate(self.tasks, 1):
                name = task.get('name', '제목 없음')
                desc = task.get('desc', '') or '설명 없음'
                path = task.get('path', '')

                # 파일 아이콘
                file_icon = self.get_file_icon(path)

                if self.file_exists(path):
                    mod_time = self.get_file_modified_time(path)
                    days = self.get_days_since_modified(path)
                    status_badge = self.get_status_badge(days)
                    name_color = Fore.WHITE
                    number_color = Fore.LIGHTYELLOW_EX
                else:
                    status_badge = self.get_status_badge(None)
                    name_color = Fore.LIGHTBLACK_EX
                    number_color = Fore.LIGHTBLACK_EX

                # 작업 카드 (더 화려하게)
                # 번호와 제목
                line1 = f"{number_color}{Style.BRIGHT}【{idx}】{Style.RESET_ALL} {file_icon} {name_color}{Style.BRIGHT}{name}{Style.RESET_ALL}"
                print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{self._card_line(line1)}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")

                # 설명
                line2 = f"     {Fore.LIGHTBLUE_EX}▸ {desc}{Style.RESET_ALL}"
                print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{self._card_line(line2)}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")

                # 경로
                path_short = path if len(path) <= UI_WIDTH - 18 else path[: UI_WIDTH - 21] + "..."
                line3 = f"     {Fore.LIGHTBLACK_EX}📂 {path_short}{Style.RESET_ALL}"
                print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{self._card_line(line3)}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")

                # 상태
                line4 = f"     {status_badge}"
                print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{self._card_line(line4)}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")

                # 구분선
                if idx < len(self.tasks):
                    separator = "─" * (UI_WIDTH - 6)
                    print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}   {Fore.LIGHTBLACK_EX}{separator}{Style.RESET_ALL}   {Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")

        # 하단 컨트롤 메뉴
        print(f"{Fore.CYAN}{Style.BRIGHT}╠{'═' * (UI_WIDTH - 2)}╣{Style.RESET_ALL}")
        print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{' ' * (UI_WIDTH - 2)}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")

        menu_parts = [
            f"{Fore.LIGHTYELLOW_EX}{Style.BRIGHT}[번호]{Style.RESET_ALL} {ICONS['rocket']} 실행",
            f"{Fore.LIGHTGREEN_EX}{Style.BRIGHT}[E]{Style.RESET_ALL} {ICONS['excel']} 엑셀편집",
            f"{Fore.LIGHTRED_EX}{Style.BRIGHT}[Q]{Style.RESET_ALL} {ICONS['target']} 종료"
        ]
        menu_text = "  │  ".join(menu_parts)
        print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{self._center_in_box(menu_text, UI_WIDTH - 2)}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")
        print(f"{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}{' ' * (UI_WIDTH - 2)}{Fore.CYAN}{Style.BRIGHT}║{Style.RESET_ALL}")
        print(f"{Fore.CYAN}{Style.BRIGHT}╚{'═' * (UI_WIDTH - 2)}╝{Style.RESET_ALL}\n")

    def clean_path(self, path_str):
        """파일 경로에서 따옴표 등 제거"""
        if not path_str:
            return ""
        # 앞뒤 공백 제거
        path_str = path_str.strip()
        # 따옴표 제거 (큰따옴표, 작은따옴표)
        if path_str.startswith('"') and path_str.endswith('"'):
            path_str = path_str[1:-1]
        if path_str.startswith("'") and path_str.endswith("'"):
            path_str = path_str[1:-1]
        return path_str.strip()

    def print_loading_animation(self, message="처리 중", duration=1.0):
        """로딩 애니메이션"""
        frames = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]
        end_time = time.time() + duration
        i = 0
        while time.time() < end_time:
            frame = frames[i % len(frames)]
            print(f"\r{Fore.CYAN}{frame} {message}...{Style.RESET_ALL}", end="", flush=True)
            time.sleep(0.1)
            i += 1
        print(f"\r{' ' * (len(message) + 10)}\r", end="")

    def edit_excel_and_reload(self):
        """엑셀 파일을 열고, 저장 후 Enter 시 목록 다시 읽기"""
        self.open_excel_for_edit()
        print(f"\n{Fore.LIGHTCYAN_EX}{'═' * 60}{Style.RESET_ALL}")
        print(f"{Fore.CYAN}{ICONS['excel']} 엑셀에서 작업 목록을 편집하세요{Style.RESET_ALL}")
        print(f"{Fore.LIGHTBLACK_EX}  1. 작업 추가/수정/삭제{Style.RESET_ALL}")
        print(f"{Fore.LIGHTBLACK_EX}  2. 파일 저장 (Ctrl+S){Style.RESET_ALL}")
        print(f"{Fore.LIGHTBLACK_EX}  3. 엑셀 닫기{Style.RESET_ALL}")
        print(f"{Fore.LIGHTCYAN_EX}{'═' * 60}{Style.RESET_ALL}\n")
        input(f"{Fore.LIGHTYELLOW_EX}▸ 작업 완료 후 Enter를 누르세요...{Style.RESET_ALL}")

        self.print_loading_animation("작업 목록 로딩", 0.8)
        self.load_tasks_from_excel()

        print(f"\n{Fore.GREEN}{ICONS['success']} 작업 목록을 업데이트했습니다!{Style.RESET_ALL}")
        print(f"{Fore.LIGHTGREEN_EX}  ▸ 총 {len(self.tasks)}개의 작업이 등록되어 있습니다{Style.RESET_ALL}\n")
        input(f"{Fore.CYAN}계속하려면 Enter를 누르세요...{Style.RESET_ALL}")

    def execute_file(self, filepath):
        """파일 실행 (개선된 버전)"""
        print(f"\n{Fore.LIGHTCYAN_EX}{'═' * 60}{Style.RESET_ALL}")

        if not os.path.exists(filepath):
            print(f"\n{Fore.RED}{ICONS['error']} 파일을 찾을 수 없습니다{Style.RESET_ALL}")
            print(f"{Fore.LIGHTRED_EX}  ▸ {filepath}{Style.RESET_ALL}\n")
            print(f"{Fore.LIGHTCYAN_EX}{'═' * 60}{Style.RESET_ALL}\n")
            input(f"{Fore.CYAN}계속하려면 Enter를 누르세요...{Style.RESET_ALL}")
            return

        try:
            file_ext = Path(filepath).suffix.lower()
            file_name = Path(filepath).name

            # 파일 정보 표시
            print(f"\n{Fore.LIGHTYELLOW_EX}{ICONS['lightning']} 작업 실행{Style.RESET_ALL}")
            print(f"{Fore.LIGHTBLUE_EX}  파일: {file_name}{Style.RESET_ALL}")
            print(f"{Fore.LIGHTBLACK_EX}  경로: {filepath}{Style.RESET_ALL}\n")

            # 로딩 애니메이션
            self.print_loading_animation("실행 준비", 0.5)

            if file_ext == '.py':
                # Python 파일 실행
                print(f"{Fore.GREEN}{ICONS['python']} Python 스크립트를 실행합니다...{Style.RESET_ALL}")
                subprocess.Popen([sys.executable, filepath],
                               creationflags=subprocess.CREATE_NEW_CONSOLE if os.name == 'nt' else 0)
            else:
                # 기타 파일 실행 (엑셀, 폴더 등)
                icon = self.get_file_icon(filepath)
                print(f"{Fore.GREEN}{icon} 파일을 실행합니다...{Style.RESET_ALL}")
                if os.name == 'nt':
                    os.startfile(filepath)
                else:
                    subprocess.Popen(['xdg-open', filepath])

            time.sleep(0.3)
            print(f"\n{Fore.LIGHTGREEN_EX}{ICONS['success']} 실행 완료!{Style.RESET_ALL}\n")
            print(f"{Fore.LIGHTCYAN_EX}{'═' * 60}{Style.RESET_ALL}\n")
            input(f"{Fore.CYAN}계속하려면 Enter를 누르세요...{Style.RESET_ALL}")

        except Exception as e:
            print(f"\n{Fore.RED}{ICONS['error']} 실행 오류 발생{Style.RESET_ALL}")
            print(f"{Fore.LIGHTRED_EX}  ▸ {str(e)}{Style.RESET_ALL}\n")
            print(f"{Fore.LIGHTCYAN_EX}{'═' * 60}{Style.RESET_ALL}\n")
            input(f"{Fore.CYAN}계속하려면 Enter를 누르세요...{Style.RESET_ALL}")

    def show_startup_animation(self):
        """시작 애니메이션"""
        self.clear_screen()
        print("\n" * 5)

        # 로고 애니메이션
        logo = [
            "    ╔═══════════════════════════════════════╗",
            "    ║                                       ║",
            "    ║   🚀  JP 통합업무 대시보드  ⭐      ║",
            "    ║                                       ║",
            "    ║      Task Launcher v2.0               ║",
            "    ║                                       ║",
            "    ╚═══════════════════════════════════════╝",
        ]

        for line in logo:
            print(f"{Fore.CYAN}{Style.BRIGHT}{line}{Style.RESET_ALL}")
            time.sleep(0.1)

        print("\n")
        self.print_loading_animation("시스템 초기화", 1.2)
        print(f"{Fore.GREEN}{ICONS['success']} 준비 완료!\n{Style.RESET_ALL}")
        time.sleep(0.5)

    def run(self):
        """메인 루프: 창 열림 → 메뉴 표시. 수정 시 [E]로 엑셀 열어 편집 후 다시 읽기"""
        if os.name == 'nt':
            os.system('title 🚀 JP 통합업무 대시보드')
            setup_console_appearance()

        # 시작 애니메이션
        self.show_startup_animation()

        self.ensure_excel_template()
        self.load_tasks_from_excel()

        while True:
            self.display_main_menu()

            choice = input(f"{Fore.LIGHTYELLOW_EX}{Style.BRIGHT}▸ 선택: {Style.RESET_ALL}").strip().upper()

            if choice == 'Q':
                # 종료 애니메이션
                print(f"\n{Fore.LIGHTCYAN_EX}{'═' * 60}{Style.RESET_ALL}")
                print(f"\n{Fore.LIGHTMAGENTA_EX}{ICONS['gem']} 프로그램을 종료합니다...{Style.RESET_ALL}\n")
                self.print_loading_animation("종료 중", 0.8)
                print(f"\n{Fore.MAGENTA}{ICONS['star']} 이용해 주셔서 감사합니다!{Style.RESET_ALL}\n")
                print(f"{Fore.LIGHTCYAN_EX}{'═' * 60}{Style.RESET_ALL}\n")
                time.sleep(0.5)
                break

            elif choice == 'E':
                self.edit_excel_and_reload()

            elif choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(self.tasks):
                    task = self.tasks[idx]
                    self.execute_file(task.get('path', ''))
                else:
                    print(f"\n{Fore.RED}{ICONS['error']} 올바른 번호를 입력하세요 (1-{len(self.tasks)}){Style.RESET_ALL}")
                    input(f"\n{Fore.CYAN}계속하려면 Enter를 누르세요...{Style.RESET_ALL}")

            else:
                print(f"\n{Fore.RED}{ICONS['warning']} 올바른 옵션을 선택하세요 (번호/E/Q){Style.RESET_ALL}")
                input(f"\n{Fore.CYAN}계속하려면 Enter를 누르세요...{Style.RESET_ALL}")


def launch_in_new_console():
    """
    Windows: 새 독립 콘솔 창을 띄우고, 그 창 안에서만 이 프로그램이 실행되게 함.
    IDE 터미널에서는 바로 종료되고, 새 창에서만 실행됨.
    """
    if os.name != 'nt':
        return False
    if "--in-console" in sys.argv:
        return False  # 이미 새 창에서 실행 중
    script_path = Path(__file__).resolve()
    try:
        # CREATE_NEW_CONSOLE: IDE 터미널이 아닌 새 콘솔 창에서만 실행
        subprocess.Popen(
            [sys.executable, "-u", str(script_path), "--in-console"],
            cwd=str(script_path.parent),
            creationflags=subprocess.CREATE_NEW_CONSOLE,
        )
        return True
    except Exception:
        return False


if __name__ == "__main__":
    if launch_in_new_console():
        sys.exit(0)
    launcher = TaskLauncher()
    launcher.run()
