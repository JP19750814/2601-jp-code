import os
import platform
import subprocess
from openpyxl import Workbook, load_workbook

def collect_subfolders(folder_path):
    """
    folder_path 및 모든 하위 폴더를 순회하여,
    디렉터리 경로만 리스트로 수집(파일 제외).
    """
    subfolders = []
    for root, dirs, files in os.walk(folder_path):
        for d in dirs:
            full_path = os.path.join(root, d)
            subfolders.append(full_path)
    return subfolders

def remove_prefix_in_excel(file_path, sheet_name, target_col, prefix_to_remove):
    """
    이미 존재하는 엑셀 파일(file_path)을 열어,
    sheet_name 시트의 target_col 열에서 prefix_to_remove로 시작하는 접두어를 제거.
    예) 'M:\\애니노블\\폴더이름' → '폴더이름'
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    for cell in ws[target_col]:
        if cell.value and isinstance(cell.value, str):
            if cell.value.startswith(prefix_to_remove):
                # 접두어 길이만큼 잘라내고 나머지 부분만 유지
                cell.value = cell.value[len(prefix_to_remove):]

    wb.save(file_path)

def open_excel_file(file_path):
    """
    운영체제별로 엑셀(또는 기본 프로그램)로 파일을 여는 함수.
    """
    system_name = platform.system()
    if system_name == "Windows":
        os.startfile(file_path)  # Windows
    elif system_name == "Darwin":
        subprocess.run(["open", file_path])  # Mac
    else:
        subprocess.run(["xdg-open", file_path])  # Linux

def export_subfolders_and_remove_prefix(folder_path, excel_path, prefix_to_remove):
    """
    1) folder_path의 모든 하위 폴더를 찾아 엑셀에 기록.
    2) 접두어(prefix_to_remove)가 있으면, 엑셀에서 해당 접두어를 제거.
    3) 엑셀 파일을 자동으로 열기.
    """
    # 1) 하위 폴더 수집
    subfolders = collect_subfolders(folder_path)

    # 2) 새 엑셀 만들기
    wb = Workbook()
    ws = wb.active
    ws.title = "Subfolders"

    # 헤더(원하시는 대로 수정 가능)
    ws.append(["폴더 경로"])

    # 폴더 경로를 A열에 기록
    for sf in subfolders:
        ws.append([sf])

    # 저장
    wb.save(excel_path)
    print(f"[완료] 엑셀 파일로 저장: {excel_path}")

    # 3) 접두어 제거 (prefix_to_remove가 빈 문자열이면 생략)
    if prefix_to_remove:
        remove_prefix_in_excel(file_path=excel_path,
                               sheet_name="Subfolders",
                               target_col="A",
                               prefix_to_remove=prefix_to_remove)
        print(f"[완료] 접두어 '{prefix_to_remove}' 제거")

    # 4) 엑셀 파일 열기
    open_excel_file(excel_path)

def main():
    # 조사할 폴더 경로 입력
    folder_path = input("조사할 폴더 경로를 입력하세요: ").strip()

    # 엑셀 파일명 입력 (.xlsx 자동 추가)
    raw_excel_name = input("저장할 엑셀 파일명을 입력하세요 (.xlsx 생략 가능): ").strip()
    if not raw_excel_name.lower().endswith(".xlsx"):
        raw_excel_name += ".xlsx"

    # 제거할 접두어 입력 (예: M:\애니노블\)
    prefix = input("제거할 접두어가 있으면 입력하세요 (예: M:\\애니노블\\). 없으면 엔터: ").strip()

    # 전체 작업 실행
    export_subfolders_and_remove_prefix(folder_path, raw_excel_name, prefix)

if __name__ == "__main__":
    main()
